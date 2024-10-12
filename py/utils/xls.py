# Licensed under the Apache License, Version 2.0 (the "License");
# Copyright 2024 Øivind Loe
# See LICENSE file or http://www.apache.org/licenses/LICENSE-2.0 for details.
# ~

import datetime
import logging
import warnings
from contextlib import contextmanager

import pandas as pd

from olib.py.utils.csv import iterCSV
from olib.py.utils.date import defaultTimezone, fileModifiedTime
from olib.py.utils.file import acceptableFilename

logger = logging.getLogger(__name__)


@contextmanager
def openXLS(filenameOrFile, storage=None):
    from openpyxl import load_workbook  # Deferred for speed

    # Expect same format as is written in 'writeXls'
    logger.info('loading: ' + filenameOrFile if isinstance(filenameOrFile, str) else '<bytes>')
    with warnings.catch_warnings():
        # DO NOT TURN ON READ ONLY for openpyxl as it makes reads extremely slow
        # Catch warning about openpyxl not understanding an excel extension
        warnings.simplefilter('ignore')
        if storage is not None:
            # This version supports remote storage (S3)
            with storage.open(filenameOrFile, 'rb') as f:
                wb = load_workbook(f, data_only=True, keep_vba=False, keep_links=False)
                fileMod = storage.get_modified_time(filenameOrFile)
        else:
            # This version only supports local files, e.g. for unittests
            wb = load_workbook(filenameOrFile, data_only=True, keep_vba=False, keep_links=False)
            if isinstance(filenameOrFile, str):
                fileMod = defaultTimezone(datetime.datetime.fromtimestamp(fileModifiedTime(filenameOrFile)))
            else:
                fileMod = None

    try:
        yield wb, fileMod
    finally:
        wb.close()


@contextmanager
def readXLS(
    filenameOrFile,
    sheet: str | None = None,
    headerRowFirst=None,
    skipRows=0,
    storage=None,
    yieldIterator=False,
):
    """
    Read a single-sheet excel file
    :param sheet: if not specified, the first sheet is selected
    """
    with openXLS(filenameOrFile, storage=storage) as (wb, _):
        if sheet is None:
            sheet = wb.sheetnames[0]

        # Find columns from first row
        iterator = wb[sheet].iter_rows()

        if yieldIterator:
            yield iterator
        else:
            # Shared implementation with CSV files
            yield iterCSV(iterator, headerRowFirst, skipRows)


def writeXLS(filename, sheet, data, raw=False):
    """
    Write any data acceptable by pd.DataFrame(...) to XLS
    :param raw: If true, data bypasses dataframe, and instead a double array is expected
    """
    from openpyxl import Workbook  # Deferred for startup speed

    wb = Workbook(write_only=True)
    # Need to truncate title (title on the tab) to 31 to make sure Excel doesn't try to repair it on windows
    ws = wb.create_sheet(title=acceptableFilename(sheet)[:31])

    if not raw:
        df = pd.DataFrame(data)
        data = [df.columns.values] + [tuple(r) for rowI, r in df.iterrows()]

    for l in data:
        ws.append(list(l))

    wb.save(filename)
